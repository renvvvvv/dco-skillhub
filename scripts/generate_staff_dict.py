"""从 Excel 生成人员字典 JSON"""
import json
from pathlib import Path
from openpyxl import load_workbook

EXCEL_PATH = Path(r"C:\Users\wuton\Desktop\ceshi-3\全国运维人员信息台账_1. 420运维人员信息表（每日刷新）(1).xlsx")
OUTPUT_PATH = Path(__file__).parent.parent / "frontend" / "public" / "staff_dictionary.json"

def main():
    print(f"读取 Excel: {EXCEL_PATH}")
    
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    
    # 查找表头行
    headers = []
    header_row = 1
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        if any(cell for cell in row if cell):
            headers = [str(cell).strip() if cell else "" for cell in row]
            header_row = ws.iter_rows(min_row=1, max_row=5, values_only=True).__next__()
            break
    
    # 重新获取表头
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        if any(cell for cell in row if cell):
            headers = [str(cell).strip() if cell else "" for cell in row]
            header_row_idx = idx
            break
    
    print(f"表头行: {header_row_idx}")
    print(f"表头: {headers}")
    
    # 查找字段索引
    def find_col_index(keywords):
        for i, h in enumerate(headers):
            h_lower = h.lower()
            for kw in keywords:
                if kw in h_lower:
                    return i
        return None
    
    name_idx = find_col_index(["姓名", "name"])
    new_id_idx = find_col_index(["新员工编号", "新工号"])
    old_id_idx = find_col_index(["员工编号", "工号"])
    dept_idx = find_col_index(["部门", "department"])
    org_idx = find_col_index(["组织", "公司", "organization", "单位"])
    
    print(f"字段索引 - 姓名:{name_idx}, 新员工编号:{new_id_idx}, 员工编号:{old_id_idx}, 部门:{dept_idx}, 组织:{org_idx}")
    
    if name_idx is None:
        print("错误: 未找到姓名字段")
        return
    
    # 提取数据
    staff_list = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        name = row[name_idx] if name_idx is not None else None
        if not name or str(name).strip() == "":
            continue
        
        staff = {
            "name": str(name).strip(),
            "employee_id": str(row[old_id_idx]).strip() if old_id_idx is not None and row[old_id_idx] else "",
            "new_employee_id": str(row[new_id_idx]).strip() if new_id_idx is not None and row[new_id_idx] else "",
            "department": str(row[dept_idx]).strip() if dept_idx is not None and row[dept_idx] else "",
            "organization": str(row[org_idx]).strip() if org_idx is not None and row[org_idx] else ""
        }
        staff_list.append(staff)
    
    print(f"提取到 {len(staff_list)} 条人员记录")
    
    # 保存 JSON
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(staff_list, f, ensure_ascii=False, indent=2)
    
    print(f"字典已保存: {OUTPUT_PATH}")

if __name__ == "__main__":
    main()
